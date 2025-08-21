import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.chart import BarChart, Reference, PieChart
from openpyxl.utils.dataframe import dataframe_to_rows
from datetime import datetime
import os

class ReportAgent:
    def __init__(self):
        self.reports_dir = "reports"
        self._ensure_reports_directory()
    
    def _ensure_reports_directory(self):
        """確保報告目錄存在"""
        if not os.path.exists(self.reports_dir):
            os.makedirs(self.reports_dir)
    
    def generate_excel_report(self, returns_data, statistics_data=None, report_type="comprehensive"):
        """生成 Excel 報告"""
        try:
            print(f"開始生成 Excel 報告，資料類型: {type(returns_data)}")
            print(f"資料長度: {len(returns_data) if returns_data is not None else 'None'}")
            
            # 確保報告目錄存在
            self._ensure_reports_directory()
            
            # 建立工作簿
            wb = Workbook()
            
            # 移除預設工作表
            wb.remove(wb.active)
            
            # 建立摘要工作表
            summary_sheet = wb.create_sheet("摘要")
            self._create_summary_sheet(summary_sheet, returns_data, statistics_data)
            
            # 建立詳細資料工作表
            details_sheet = wb.create_sheet("詳細資料")
            self._create_details_sheet(details_sheet, returns_data)
            
            # 建立分析工作表
            analysis_sheet = wb.create_sheet("分析")
            self._create_analysis_sheet(analysis_sheet, statistics_data)
            
            # 建立發現工作表
            findings_sheet = wb.create_sheet("發現")
            self._create_findings_sheet(findings_sheet, returns_data, statistics_data)
            
            # 生成檔案名稱
            timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
            filename = f"returns_report_{timestamp}.xlsx"
            filepath = os.path.join(self.reports_dir, filename)
            
            print(f"儲存報告到: {filepath}")
            # 儲存檔案
            wb.save(filepath)
            print("報告儲存成功")
            
            return {
                'status': 'success',
                'message': f'Excel 報告生成成功',
                'data': {
                    'filename': filename,
                    'filepath': filepath,
                    'sheets': ['摘要', '詳細資料', '分析', '發現']
                }
            }
            
        except Exception as e:
            print(f"生成 Excel 報告時發生錯誤: {e}")
            import traceback
            traceback.print_exc()
            return {
                'status': 'error',
                'message': f'生成 Excel 報告時發生錯誤: {str(e)}'
            }
    
    def _create_summary_sheet(self, sheet, returns_data, statistics_data):
        """建立摘要工作表"""
        # 設定欄寬
        sheet.column_dimensions['A'].width = 20
        sheet.column_dimensions['B'].width = 30
        
        # 標題樣式
        title_font = Font(bold=True, size=16)
        header_font = Font(bold=True, size=12)
        
        # 標題
        sheet['A1'] = "退貨分析報告摘要"
        sheet['A1'].font = title_font
        sheet.merge_cells('A1:B1')
        
        # 報告資訊
        current_time = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
        sheet['A3'] = "報告生成時間"
        sheet['B3'] = current_time
        sheet['A3'].font = header_font
        
        sheet['A4'] = "總退貨記錄數"
        sheet['B4'] = len(returns_data) if returns_data else 0
        sheet['A4'].font = header_font
        
        if statistics_data:
            sheet['A5'] = "總退貨數量"
            sheet['B5'] = statistics_data.get('total_returns', 0)
            sheet['A5'].font = header_font
            
            sheet['A6'] = "涉及商店數量"
            sheet['B6'] = len(statistics_data.get('store_stats', []))
            sheet['A6'].font = header_font
            
            sheet['A7'] = "涉及產品數量"
            sheet['B7'] = len(statistics_data.get('product_stats', []))
            sheet['A7'].font = header_font
        
        # 樣式設定
        for row in range(3, 8):
            sheet[f'A{row}'].font = Font(bold=True)
            sheet[f'B{row}'].font = Font(size=11)
    
    def _create_details_sheet(self, sheet, returns_data):
        """建立詳細資料工作表"""
        if not returns_data or len(returns_data) == 0:
            sheet['A1'] = "無退貨資料"
            return
        
        # 設定欄寬
        sheet.column_dimensions['A'].width = 15
        sheet.column_dimensions['B'].width = 20
        sheet.column_dimensions['C'].width = 25
        sheet.column_dimensions['D'].width = 15
        sheet.column_dimensions['E'].width = 20
        sheet.column_dimensions['F'].width = 20
        
        # 標題
        headers = ['ID', '訂單ID', '產品名稱', '商店名稱', '退貨日期', '建立時間']
        for col, header in enumerate(headers, 1):
            cell = sheet.cell(row=1, column=col, value=header)
            cell.font = Font(bold=True)
            cell.fill = PatternFill(start_color="CCCCCC", end_color="CCCCCC", fill_type="solid")
        
        # 資料
        for row_idx, record in enumerate(returns_data, 2):
            try:
                # 安全地獲取資料
                record_id = str(record.get('id', '')) if record.get('id') is not None else ''
                order_id = str(record.get('order_id', '')) if record.get('order_id') is not None else ''
                product = str(record.get('product', '')) if record.get('product') is not None else ''
                store_name = str(record.get('store_name', '')) if record.get('store_name') is not None else ''
                return_date = str(record.get('return_date', '')) if record.get('return_date') is not None else ''
                created_at = str(record.get('created_at', '')) if record.get('created_at') is not None else ''
                
                sheet.cell(row=row_idx, column=1, value=record_id)
                sheet.cell(row=row_idx, column=2, value=order_id)
                sheet.cell(row=row_idx, column=3, value=product)
                sheet.cell(row=row_idx, column=4, value=store_name)
                sheet.cell(row=row_idx, column=5, value=return_date)
                sheet.cell(row=row_idx, column=6, value=created_at)
                
            except Exception as row_error:
                print(f"處理第 {row_idx} 行資料時發生錯誤: {row_error}")
                continue
        
        # 邊框樣式
        thin_border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )
        
        for row in range(1, len(returns_data) + 2):
            for col in range(1, len(headers) + 1):
                try:
                    sheet.cell(row=row, column=col).border = thin_border
                except:
                    pass
    
    def _create_analysis_sheet(self, sheet, statistics_data):
        """建立分析工作表"""
        if not statistics_data:
            sheet['A1'] = "無統計資料"
            return
        
        # 設定欄寬
        sheet.column_dimensions['A'].width = 20
        sheet.column_dimensions['B'].width = 15
        
        # 標題
        sheet['A1'] = "退貨分析"
        sheet['A1'].font = Font(bold=True, size=14)
        
        # 商店統計
        sheet['A3'] = "按商店統計"
        sheet['A3'].font = Font(bold=True, size=12)
        
        headers = ['商店名稱', '退貨數量']
        for col, header in enumerate(headers, 1):
            cell = sheet.cell(row=4, column=col, value=header)
            cell.font = Font(bold=True)
            cell.fill = PatternFill(start_color="E6E6E6", end_color="E6E6E6", fill_type="solid")
        
        store_stats = statistics_data.get('store_stats', [])
        for row_idx, stat in enumerate(store_stats, 5):
            sheet.cell(row=row_idx, column=1, value=stat.get('store_name', ''))
            sheet.cell(row=row_idx, column=2, value=stat.get('count', 0))
        
        # 產品統計
        start_row = 5 + len(store_stats) + 2
        sheet[f'A{start_row}'] = "按產品統計"
        sheet[f'A{start_row}'].font = Font(bold=True, size=12)
        
        for col, header in enumerate(headers, 1):
            cell = sheet.cell(row=start_row + 1, column=col, value=header)
            cell.font = Font(bold=True)
            cell.fill = PatternFill(start_color="E6E6E6", end_color="E6E6E6", fill_type="solid")
        
        product_stats = statistics_data.get('product_stats', [])
        for row_idx, stat in enumerate(product_stats, start_row + 2):
            sheet.cell(row=row_idx, column=1, value=stat.get('product', ''))
            sheet.cell(row=row_idx, column=2, value=stat.get('count', 0))
        
        # 月份統計
        month_start_row = start_row + 2 + len(product_stats) + 2
        sheet[f'A{month_start_row}'] = "按月份統計"
        sheet[f'A{month_start_row}'].font = Font(bold=True, size=12)
        
        month_headers = ['月份', '退貨數量']
        for col, header in enumerate(month_headers, 1):
            cell = sheet.cell(row=month_start_row + 1, column=col, value=header)
            cell.font = Font(bold=True)
            cell.fill = PatternFill(start_color="E6E6E6", end_color="E6E6E6", fill_type="solid")
        
        monthly_stats = statistics_data.get('monthly_stats', [])
        for row_idx, stat in enumerate(monthly_stats, month_start_row + 2):
            sheet.cell(row=row_idx, column=1, value=stat.get('month', ''))
            sheet.cell(row=row_idx, column=2, value=stat.get('count', 0))
    
    def _create_findings_sheet(self, sheet, returns_data, statistics_data):
        """建立發現工作表"""
        # 設定欄寬
        sheet.column_dimensions['A'].width = 20
        sheet.column_dimensions['B'].width = 50
        
        # 標題
        sheet['A1'] = "主要發現"
        sheet['A1'].font = Font(bold=True, size=14)
        
        findings = []
        
        if returns_data:
            findings.append("退貨記錄總覽")
            findings.append(f"系統中共有 {len(returns_data)} 筆退貨記錄")
            
            if statistics_data:
                total_returns = statistics_data.get('total_returns', 0)
                findings.append(f"總退貨數量: {total_returns}")
                
                store_stats = statistics_data.get('store_stats', [])
                if store_stats:
                    top_store = store_stats[0]
                    findings.append(f"退貨最多的商店: {top_store.get('store_name', '')} ({top_store.get('count', 0)} 筆)")
                
                product_stats = statistics_data.get('product_stats', [])
                if product_stats:
                    top_product = product_stats[0]
                    findings.append(f"退貨最多的產品: {top_product.get('product', '')} ({top_product.get('count', 0)} 筆)")
                
                monthly_stats = statistics_data.get('monthly_stats', [])
                if monthly_stats:
                    recent_month = monthly_stats[0]
                    findings.append(f"最近月份退貨數量: {recent_month.get('month', '')} ({recent_month.get('count', 0)} 筆)")
        
        # 寫入發現內容
        for row_idx, finding in enumerate(findings, 3):
            if row_idx == 3:  # 第一個發現
                sheet.cell(row=row_idx, column=1, value="發現項目").font = Font(bold=True)
                sheet.cell(row=row_idx, column=2, value="內容").font = Font(bold=True)
                row_idx += 1
            
            sheet.cell(row=row_idx, column=1, value=f"發現 {row_idx - 3}")
            sheet.cell(row=row_idx, column=2, value=finding)
        
        # 建議
        suggestions_start_row = len(findings) + 5
        sheet[f'A{suggestions_start_row}'] = "建議"
        sheet[f'A{suggestions_start_row}'].font = Font(bold=True, size=12)
        
        suggestions = [
            "定期分析退貨趨勢，識別問題產品",
            "關注高退貨率商店，提供支援和培訓",
            "分析退貨原因，改善產品品質",
            "建立預警機制，及時發現異常退貨"
        ]
        
        for idx, suggestion in enumerate(suggestions, 1):
            sheet.cell(row=suggestions_start_row + idx, column=1, value=f"建議 {idx}")
            sheet.cell(row=suggestions_start_row + idx, column=2, value=suggestion)
    
    def generate_simple_report(self, returns_data):
        """生成簡單報告（用於快速測試）"""
        try:
            print(f"開始生成簡單報告，資料類型: {type(returns_data)}")
            print(f"資料長度: {len(returns_data) if returns_data is not None else 'None'}")
            
            # 確保報告目錄存在
            self._ensure_reports_directory()
            
            wb = Workbook()
            ws = wb.active
            ws.title = "退貨記錄"
            
            # 標題
            ws['A1'] = "退貨記錄報告"
            ws['A1'].font = Font(bold=True, size=16)
            
            # 檢查資料
            if returns_data is not None and len(returns_data) > 0:
                print(f"處理 {len(returns_data)} 筆記錄")
                
                # 標題行
                headers = ['訂單ID', '產品', '商店名稱', '日期']
                for col, header in enumerate(headers, 1):
                    cell = ws.cell(row=3, column=col, value=header)
                    cell.font = Font(bold=True)
                
                # 資料
                for row_idx, record in enumerate(returns_data, 4):
                    try:
                        # 安全地獲取資料
                        order_id = str(record.get('order_id', '')) if record.get('order_id') is not None else ''
                        product = str(record.get('product', '')) if record.get('product') is not None else ''
                        store_name = str(record.get('store_name', '')) if record.get('store_name') is not None else ''
                        return_date = str(record.get('return_date', '')) if record.get('return_date') is not None else ''
                        
                        ws.cell(row=row_idx, column=1, value=order_id)
                        ws.cell(row=row_idx, column=2, value=product)
                        ws.cell(row=row_idx, column=3, value=store_name)
                        ws.cell(row=row_idx, column=4, value=return_date)
                        
                    except Exception as row_error:
                        print(f"處理第 {row_idx} 行資料時發生錯誤: {row_error}")
                        continue
            else:
                print("沒有退貨資料，建立空報告")
                ws['A3'] = "暫無退貨記錄"
            
            # 儲存
            timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
            filename = f"simple_report_{timestamp}.xlsx"
            filepath = os.path.join(self.reports_dir, filename)
            
            print(f"儲存報告到: {filepath}")
            wb.save(filepath)
            print("報告儲存成功")
            
            return {
                'status': 'success',
                'message': '簡單報告生成成功',
                'data': {
                    'filename': filename,
                    'filepath': filepath
                }
            }
            
        except Exception as e:
            print(f"生成簡單報告時發生錯誤: {e}")
            import traceback
            traceback.print_exc()
            return {
                'status': 'error',
                'message': f'生成簡單報告時發生錯誤: {str(e)}'
            }
